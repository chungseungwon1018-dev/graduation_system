import pandas as pd
import mysql.connector
from mysql.connector import Error
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import re
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class XlsxParser:
    def __init__(self, db_config: Dict[str, str]):
        self.db_config = db_config
        self.connection = None

    def connect_db(self):
        try:
            self.connection = mysql.connector.connect(**self.db_config)
            logger.info("MySQL 데이터베이스 연결 성공")
        except Error as e:
            logger.error(f"MySQL 연결 오류: {e}")
            raise

    def disconnect_db(self):
        if self.connection and self.connection.is_connected():
            self.connection.close()
            logger.info("MySQL 연결 해제")

    def parse_excel_file(self, file_path: str) -> Tuple[Dict, List[Dict]]:
        try:
            try:
                logger.info("openpyxl을 직접 사용해서 파일 읽기 시도")
                return self._parse_with_openpyxl_direct(file_path)
            except Exception as openpyxl_direct_error:
                logger.warning(f"openpyxl 직접 사용 실패: {openpyxl_direct_error}")

            engines = ['openpyxl', 'xlrd']

            for engine in engines:
                try:
                    logger.info(f"{engine} 엔진으로 Excel 파일 읽기 시도")
                    if engine == 'xlrd' and file_path.endswith('.xlsx'):
                        continue

                    with pd.ExcelFile(file_path, engine=engine) as excel_file:
                        logger.info(f"{engine} 엔진으로 성공적으로 파일 열기 완료")

                        personal_info = {}
                        course_records = []

                        logger.info(f"Excel 시트 목록: {excel_file.sheet_names}")

                        for sheet_name in excel_file.sheet_names:
                            try:
                                logger.info(f"시트 '{sheet_name}' 처리 중")
                                df = excel_file.parse(sheet_name)

                                if '개인정보' in sheet_name or '학적' in sheet_name:
                                    personal_info = self._extract_personal_info(df)
                                    logger.info(f"개인정보 추출 완료: {len(personal_info)} 항목")
                                elif '성적' in sheet_name or '이수' in sheet_name:
                                    course_records = self._extract_course_records(df)
                                    logger.info(f"성적 데이터 추출 완료: {len(course_records)} 항목")
                                else:
                                    combined_data = self._try_extract_combined_data(df)
                                    if combined_data:
                                        personal_info, course_records = combined_data
                                        logger.info(f"결합 데이터 추출 완료: 개인정보 {len(personal_info)} 항목, 성적 {len(course_records)} 항목")
                            except Exception as sheet_error:
                                logger.warning(f"시트 {sheet_name} 처리 중 오류: {sheet_error}")
                                continue

                        logger.info(f"최종 결과: 개인정보 {len(personal_info)} 항목, 성적 {len(course_records)} 항목")
                        return personal_info, course_records

                except Exception as engine_error:
                    logger.warning(f"{engine} 엔진 실패: {engine_error}")
                    continue

            logger.error("모든 Excel 파싱 시도 실패")
            raise ValueError("Excel 파일을 읽을 수 없습니다.")

        except Exception as e:
            logger.error(f"Excel 파일 파싱 오류: {e}")
            raise
    
    def _extract_personal_info(self, df: pd.DataFrame) -> Dict:
        personal_info = {}
        
        field_mappings = {
            '대학': ['대학', '단과대학', '소속대학'],
            '학과': ['학과', '전공학과', '소속학과'],
            '전공': ['전공', '주전공', '제1전공'],
            '부전공': ['부전공', '제2전공'],
            '다전공': ['다전공', '복수전공'],
            '과정': ['과정', '학위과정'],
            '입학일자': ['입학일', '입학일자', '입학년월'],
            '학번': ['학번', '학생번호'],
            '성명': ['성명', '이름', '학생명'],
            '교과적용년도': ['교과적용년도', '적용년도'],
            '이수학기': ['이수학기', '학기'],
            '생년월일': ['생년월일', '출생일'],
            '학년': ['학년'],
            '평생사제상담건수': ['평생사제상담건수', '상담건수']
        }
        
        for field, possible_names in field_mappings.items():
            for name in possible_names:
                value = self._find_value_in_dataframe(df, name)
                if value:
                    personal_info[field] = value
                    break
        
        return personal_info
    
    def _extract_course_records(self, df: pd.DataFrame) -> List[Dict]:
        course_records = []
        
        column_mappings = {
            '구분': ['구분', '이수구분', '과목구분'],
            '영역': ['영역', '교과영역', '이수영역'],
            '세부영역': ['세부영역', '상세영역'],
            '년도': ['년도', '학년도', '이수년도'],
            '학기': ['학기', '이수학기'],
            '교과목번호': ['교과목번호', '과목번호', '강의번호'],
            '교과목명': ['교과목명', '과목명', '강의명'],
            '학점': ['학점', '이수학점'],
            '이수구분': ['이수구분', '구분'],
            '성적': ['성적', '평점', 'GPA']
        }
        
        normalized_columns = {}
        for std_name, possible_names in column_mappings.items():
            for col in df.columns:
                if any(name in str(col) for name in possible_names):
                    normalized_columns[std_name] = col
                    break
        
        for _, row in df.iterrows():
            if pd.notna(row.get(normalized_columns.get('교과목명', ''), None)):
                record = {}
                for std_name, col_name in normalized_columns.items():
                    record[std_name] = row.get(col_name, None)
                course_records.append(record)
        
        return course_records
    
    def _try_extract_combined_data(self, df: pd.DataFrame) -> Optional[Tuple[Dict, List[Dict]]]:
        personal_info = self._extract_personal_info(df)
        course_records = self._extract_course_records(df)
        
        if personal_info or course_records:
            return personal_info, course_records
        return None
    
    def _find_value_in_dataframe(self, df: pd.DataFrame, field_name: str) -> Optional[str]:
        for _, row in df.iterrows():
            for col in df.columns:
                if field_name in str(col):
                    value = row[col]
                    if pd.notna(value):
                        return str(value)
                
                cell_value = str(row[col]).strip() if pd.notna(row[col]) else ""
                if field_name in cell_value:
                    parts = cell_value.split(':')
                    if len(parts) > 1:
                        return parts[1].strip()
        return None
    
    def _parse_with_openpyxl_direct(self, file_path: str) -> Tuple[Dict, List[Dict]]:
        """openpyxl을 직접 사용해서 스타일 정보 없이 데이터만 읽기"""
        try:
            # data_only=True, read_only=True로 스타일 정보 무시
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            logger.warning(f"read_only 모드 실패: {e}")
            try:
                # read_only=False로 재시도
                workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            except Exception as e2:
                logger.warning(f"data_only 모드도 실패: {e2}")
                # 마지막 시도: 기본 모드
                workbook = openpyxl.load_workbook(file_path)
        
        personal_info = {}
        course_records = []
        major_required_credits = None
        major_elective_credits = None
        general_elective_credits = None
        
        logger.info(f"워크북 시트 목록: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            try:
                logger.info(f"시트 '{sheet_name}' 처리 중")
                worksheet = workbook[sheet_name]
                
                # AC22, AH22, Y22 셀 값 추출 (전공필수, 전공선택, 일반선택 이수학점)
                try:
                    ac22_value = worksheet['AC22'].value
                    ah22_value = worksheet['AH22'].value
                    y22_value = worksheet['Y22'].value
                    if ac22_value is not None:
                        major_required_credits = float(ac22_value)
                        logger.info(f"AC22(전공필수) 값 추출: {major_required_credits}")
                    if ah22_value is not None:
                        major_elective_credits = float(ah22_value)
                        logger.info(f"AH22(전공선택) 값 추출: {major_elective_credits}")
                    if y22_value is not None:
                        general_elective_credits = float(y22_value)
                        logger.info(f"Y22(일반선택) 값 추출: {general_elective_credits}")
                except Exception as cell_error:
                    logger.warning(f"AC22/AH22/Y22 셀 읽기 오류: {cell_error}")
                
                # 워크시트를 DataFrame으로 변환
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    logger.warning(f"시트 '{sheet_name}'에 데이터가 없음")
                    continue
                    
                # 첫 번째 행을 헤더로 사용
                headers = data[0] if data else []
                rows = data[1:] if len(data) > 1 else []
                
                # DataFrame 생성
                df = pd.DataFrame(rows, columns=headers)
                
                # 데이터 처리
                if '개인정보' in sheet_name or '학적' in sheet_name:
                    personal_info = self._extract_personal_info(df)
                    logger.info(f"개인정보 추출 완료: {len(personal_info)} 항목")
                elif '성적' in sheet_name or '이수' in sheet_name:
                    course_records = self._extract_course_records(df)
                    logger.info(f"성적 데이터 추출 완료: {len(course_records)} 항목")
                else:
                    combined_data = self._try_extract_combined_data(df)
                    if combined_data:
                        personal_info, course_records = combined_data
                        logger.info(f"결합 데이터 추출 완료: 개인정보 {len(personal_info)} 항목, 성적 {len(course_records)} 항목")
            except Exception as sheet_error:
                logger.warning(f"시트 '{sheet_name}' 처리 중 오류: {sheet_error}")
                continue
        
        workbook.close()
        
        # personal_info에 전공필수/전공선택/일반선택 학점 추가
        if major_required_credits is not None:
            personal_info['전공필수학점'] = major_required_credits
        if major_elective_credits is not None:
            personal_info['전공선택학점'] = major_elective_credits
        if general_elective_credits is not None:
            personal_info['일반선택학점'] = general_elective_credits
        
        logger.info(f"openpyxl 직접 사용 최종 결과: 개인정보 {len(personal_info)} 항목, 성적 {len(course_records)} 항목")
        return personal_info, course_records
    
    def save_to_database(self, student_id: str, personal_info: Dict, course_records: List[Dict]):
        try:
            cursor = self.connection.cursor()
            
            personal_info['학번'] = student_id
            self._save_personal_info(cursor, personal_info)
            self._save_course_records(cursor, student_id, course_records)
            
            self.connection.commit()
            logger.info(f"학번 {student_id} 데이터 저장 완료")
            
        except Error as e:
            self.connection.rollback()
            logger.error(f"데이터베이스 저장 오류: {e}")
            raise
        finally:
            if cursor:
                cursor.close()
    
    def _save_personal_info(self, cursor, personal_info: Dict):
        query = """
        INSERT INTO students (
            student_id, university, department, major, minor, double_major, 
            course_type, admission_date, name, curriculum_year, 
            semester, birth_date, grade, counseling_count, 
            major_required_credits, major_elective_credits, general_elective_credits, created_at, updated_at
        ) VALUES (
            %(학번)s, %(대학)s, %(학과)s, %(전공)s, %(부전공)s, %(다전공)s,
            %(과정)s, %(입학일자)s, %(성명)s, %(교과적용년도)s,
            %(이수학기)s, %(생년월일)s, %(학년)s, %(평생사제상담건수)s, 
            %(전공필수학점)s, %(전공선택학점)s, %(일반선택학점)s, NOW(), NOW()
        ) ON DUPLICATE KEY UPDATE
            university = VALUES(university),
            department = VALUES(department),
            major = VALUES(major),
            minor = VALUES(minor),
            double_major = VALUES(double_major),
            course_type = VALUES(course_type),
            admission_date = VALUES(admission_date),
            name = VALUES(name),
            curriculum_year = VALUES(curriculum_year),
            semester = VALUES(semester),
            birth_date = VALUES(birth_date),
            grade = VALUES(grade),
            counseling_count = VALUES(counseling_count),
            major_required_credits = VALUES(major_required_credits),
            major_elective_credits = VALUES(major_elective_credits),
            general_elective_credits = VALUES(general_elective_credits),
            updated_at = NOW()
        """
        
        cursor.execute(query, personal_info)
    
    def _save_course_records(self, cursor, student_id: str, course_records: List[Dict]):
        cursor.execute("DELETE FROM course_records WHERE student_id = %s", (student_id,))
        
        query = """
        INSERT INTO course_records (
            student_id, category, area, sub_area, year, semester,
            course_code, course_name, credit, completion_type, grade, created_at
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
        )
        """
        
        for record in course_records:
            values = (
                student_id,
                record.get('구분'),
                record.get('영역'),
                record.get('세부영역'),
                record.get('년도'),
                record.get('학기'),
                record.get('교과목번호'),
                record.get('교과목명'),
                record.get('학점'),
                record.get('이수구분'),
                record.get('성적')
            )
            cursor.execute(query, values)

def process_excel_file(file_path: str, student_id: str, db_config: Dict[str, str]) -> bool:
    parser = XlsxParser(db_config)
    
    try:
        parser.connect_db()
        personal_info, course_records = parser.parse_excel_file(file_path)
        
        if not personal_info and not course_records:
            logger.warning("Excel 파일에서 유효한 데이터를 찾을 수 없습니다.")
            return False
        
        parser.save_to_database(student_id, personal_info, course_records)
        return True
        
    except Exception as e:
        logger.error(f"Excel 파일 처리 중 오류 발생: {e}")
        return False
    finally:
        parser.disconnect_db()

if __name__ == "__main__":
    db_config = {
        'host': '203.255.78.58',
        'port': 9003,
        'database': 'graduation_system',
        'user': 'user29',
        'password': '123'
    }
    
    test_file_path = 'ex_1.xlsx'
    test_student_id = '2023001'
    
    success = process_excel_file(test_file_path, test_student_id, db_config)
    if success:
        print("Excel 파일 처리 완료")
    else:
        print("Excel 파일 처리 실패")