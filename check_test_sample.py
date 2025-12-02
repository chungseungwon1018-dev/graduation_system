import openpyxl

# test_sample_1.xlsx 파일 열기
wb = openpyxl.load_workbook('test_sample_1.xlsx')
ws = wb.active

print("=== 중요 셀 값 확인 ===")
print(f"AC22 (전공필수): {ws['AC22'].value}")
print(f"AH22 (전공선택): {ws['AH22'].value}")

# 일반선택 관련 셀 찾기
print("\n=== '일반선택' 문자열이 포함된 셀 찾기 ===")
for row in range(1, 30):
    for col in range(1, 50):
        cell = ws.cell(row=row, column=col)
        if cell.value and '일반선택' in str(cell.value):
            cell_address = cell.coordinate
            print(f"{cell_address}: {cell.value}")
            # I29 주변 5개 열 확인
            print(f"  I29 행의 값들:")
            for c in range(9, 20):  # I부터 S까지
                cell_val = ws.cell(row=29, column=c)
                print(f"    {cell_val.coordinate}: {cell_val.value}")

# 과목 데이터에서 일반선택 확인 (보통 32행부터 과목 데이터)
print("\n=== 과목 데이터에서 '일반선택' 검색 (32행부터) ===")
for row in range(32, 100):
    row_data = []
    has_data = False
    for col in range(1, 20):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            has_data = True
            if '일반선택' in str(cell.value):
                print(f"Row {row}: {[ws.cell(row=row, column=c).value for c in range(1, 20)]}")
                break
    if not has_data:
        break

wb.close()
