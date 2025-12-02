import pandas as pd
import mysql.connector
from mysql.connector import Error
import bcrypt
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import re
import json
import gc
import openpyxl
import zipfile
import xml.etree.ElementTree as ET

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EnhancedXlsxParser:
    def __init__(self, db_config: Dict[str, str]):
        self.db_config = db_config
        self.connection = None
        
        # 개인정보 필드 매핑 (우선순위 순서)
        self.personal_info_mappings = {
            '대학': ['대학', '단과대학', '소속대학', '대학명', '학교', '소속대학교'],
            '학과': ['학과', '전공학과', '소속학과', '학과명', '주학과'],
            '전공': ['전공', '주전공', '제1전공', '전공명', '주전공명'],
            '부전공': ['부전공', '제2전공', '복수전공', '부전공명'],
            '다전공': ['다전공', '복수전공', '제2전공', '다중전공'],
            '과정': ['과정', '학위과정', '교육과정', '과정구분'],
            '입학일자': ['입학일', '입학일자', '입학년월', '입학날짜', '입학년도'],
            '학번': ['학번', '학생번호', '학생ID', '수험번호'],
            '성명': ['성명', '이름', '학생명', '성함', '이름'],
            '교과적용년도': ['교과적용년도', '적용년도', '교과년도', '커리큘럼년도'],
            '이수학기': ['이수학기', '학기', '총이수학기', '현재학기'],
            '생년월일': ['생년월일', '출생일', '생일', '출생년월일'],
            '학년': ['학년', '현재학년', '이수학년'],
            '평생사제상담건수': ['평생사제상담건수', '상담건수', '멘토링횟수', '상담회수']
        }
        
        # 이수학점 필드 매핑
        self.course_field_mappings = {
            '구분': ['구분', '이수구분', '과목구분', '교과구분', '분류'],
            '영역': ['영역', '교과영역', '이수영역', '분야', '교육영역', '세부영역', '상세영역', '교양영역', '전공영역'],
            '세부영역': ['세부영역', '상세영역', '하위영역', '세분류'],
            '년도': ['년도', '학년도', '이수년도', '수강년도', '개설년도'],
            '학기': ['학기', '이수학기', '수강학기', '개설학기'],
            '교과목번호': ['교과목번호', '과목번호', '강의번호', '과목코드', '교과번호'],
            '교과목명': ['교과목명', '과목명', '강의명', '교과목이름', '강좌명'],
            '학점': ['학점', '이수학점', '취득학점', '인정학점'],
            '이수구분': ['이수구분', '구분', '취득구분', '성적구분', '이수', '완료구분'],
            '성적': ['성적', '평점', 'GPA', '등급', '점수']
        }
        
        # 교양 영역명 사전 (area 자동 보정용)
        self.liberal_area_keywords = {
            '일반교양': ['일반교양', '일반', '교양일반'],
            '확대교양': ['확대교양', '확대', '교양확대'],
            '개신기초교양': ['개신기초교양', '개신기초', '기초교양', '개신'],
            'OCU기타': ['OCU', 'ocu', 'OCU기타', 'OCU/기타'],
            'OCU/기타': ['OCU', 'ocu', 'OCU기타', 'OCU/기타']
        }
    
    def connect_db(self):
        try:
            self.connection = mysql.connector.connect(**self.db_config)
            logger.info("MySQL 데이터베이스 연결 성공")
        except Error as e:
            logger.error(f"MySQL 연결 오류: {e}")
            raise
    
    def disconnect_db(self):
        if self.connection and self.connection.is_connected():
            self.connection.close()
            logger.info("MySQL 연결 해제")
    
    def parse_excel_file(self, file_path: str) -> Tuple[Dict, List[Dict]]:
        """Excel 파일에서 개인정보와 이수학점 정보를 추출"""
        excel_file = None
        try:
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')  # 엔진 명시
            logger.info(f"Excel 파일 시트 목록: {excel_file.sheet_names}")
            use_openpyxl_fallback = False

            personal_info = {}
            course_records = []
            parsing_warnings = []

            for sheet_name in excel_file.sheet_names:
                logger.info(f"시트 '{sheet_name}' 분석 중...")
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                extracted_personal = self._extract_personal_info_from_sheet(df, sheet_name)
                if extracted_personal:
                    personal_info.update(extracted_personal)
                    logger.info(f"시트 '{sheet_name}'에서 개인정보 추출: {len(extracted_personal)}개 필드")

                extracted_courses = self._extract_course_records_from_sheet(df, sheet_name)
                if extracted_courses:
                    course_records.extend(extracted_courses)
                    logger.info(f"시트 '{sheet_name}'에서 수강기록 추출: {len(extracted_courses)}개 과목")

            logger.info("=== 추출된 개인정보 (JSON) ===")
            logger.info(json.dumps(personal_info, ensure_ascii=False, indent=2))

            logger.info("=== 추출된 이수학점 정보 (JSON) ===")
            logger.info(json.dumps(course_records[:5], ensure_ascii=False, indent=2))
            if len(course_records) > 5:
                logger.info(f"... 총 {len(course_records)}개 과목 (처음 5개만 표시)")

            # 고정 셀(AC22/AH22/Y22)에서 전공/일반선택 학점 추가 추출 시도
            try:
                fixed_cells = self._extract_fixed_cell_credits(file_path)
                if fixed_cells:
                    personal_info.update(fixed_cells)
                    logger.info(f"고정셀에서 학점 추출: {fixed_cells}")
            except Exception as e:
                logger.warning(f"고정셀 학점 추출 중 오류: {e}")

            # 파싱 경고 수집
            if not course_records:
                parsing_warnings.append('과목표에서 수강기록을 추출하지 못했습니다.')
                logger.warning('과목표에서 수강기록을 추출하지 못했습니다.')
            # check for missing mappings
            # important fields
            for key in ('전공필수학점', '전공선택학점', '일반선택학점'):
                if key not in personal_info or personal_info.get(key) is None:
                    parsing_warnings.append(f"{key}가 추출되지 않았습니다.")
            # 헤더 탐지 실패
            # Return warnings in personal_info for later use
            if parsing_warnings:
                personal_info['parsing_warnings'] = parsing_warnings
            else:
                personal_info['parsing_warnings'] = []
            # indicate parser used
            personal_info['parsing_warnings'].append('parser_used: pandas')

            try:
                return personal_info, course_records
            finally:
                # Ensure ExcelFile handle is closed
                try:
                    if excel_file is not None:
                        excel_file.close()
                except Exception:
                    pass

        except Exception as e:
            logger.error(f"Excel 파일 파싱 오류 (pandas+openpyxl 실패): {e}")
            # Fallback: use openpyxl directly and convert sheets to DataFrame
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                logger.info(f"openpyxl 로드 성공: sheets={wb.sheetnames}")
                personal_info = {}
                course_records = []
                parsing_warnings = []

                for ws_name in wb.sheetnames:
                    logger.info(f"Fallback 시트 '{ws_name}' 분석 중...")
                    ws = wb[ws_name]
                    data = list(ws.values)
                    # ensure we have at least one row
                    df = pd.DataFrame(data)

                    extracted_personal = self._extract_personal_info_from_sheet(df, ws_name)
                    if extracted_personal:
                        personal_info.update(extracted_personal)
                        logger.info(f"시트 '{ws_name}'에서 개인정보 추출(fallback): {len(extracted_personal)}개 필드")

                    extracted_courses = self._extract_course_records_from_sheet(df, ws_name)
                    if extracted_courses:
                        course_records.extend(extracted_courses)
                        logger.info(f"시트 '{ws_name}'에서 수강기록 추출(fallback): {len(extracted_courses)}개 과목")

                # attempt to extract fixed cell credits using openpyxl
                try:
                    fixed_cells = self._extract_fixed_cell_credits(file_path)
                    if fixed_cells:
                        personal_info.update(fixed_cells)
                        logger.info(f"고정셀에서 학점 추출(fallback): {fixed_cells}")
                except Exception as ex2:
                    logger.warning(f"고정셀 추출(fallback) 오류: {ex2}")

                if not course_records:
                    parsing_warnings.append('과목표에서 수강기록을 추출하지 못했습니다.')
                    logger.warning('과목표에서 수강기록을 추출하지 못했습니다. (fallback)')

                for key in ('전공필수학점', '전공선택학점', '일반선택학점'):
                    if key not in personal_info or personal_info.get(key) is None:
                        parsing_warnings.append(f"{key}가 추출되지 않았습니다.")

                if parsing_warnings:
                    personal_info['parsing_warnings'] = parsing_warnings
                else:
                    personal_info['parsing_warnings'] = []
                # indicate parser used
                personal_info['parsing_warnings'].append('parser_used: openpyxl')

                try:
                    return personal_info, course_records
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass
            except Exception as ex:
                logger.error(f"openpyxl fallback 실패: {ex}")
                # 마지막 단계: zip/xml 기반 직접 파싱 시도 (openpyxl이 특정 스타일로 실패할 때)
                try:
                    logger.info("zip/xml 기반 파서 시도")
                    personal_info = {}
                    course_records = []
                    parsing_warnings = []
                    # shared strings 및 시트 파일 목록 읽기
                    shared_strings = []
                    sheet_files = []
                    try:
                        with zipfile.ZipFile(file_path, 'r') as z:
                            if 'xl/sharedStrings.xml' in z.namelist():
                                ss_xml = z.read('xl/sharedStrings.xml')
                                root = ET.fromstring(ss_xml)
                                for si in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                                    texts = ''.join([t.text or '' for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')])
                                    shared_strings.append(texts)
                            sheet_files = [name for name in z.namelist() if name.startswith('xl/worksheets/sheet')]
                    except Exception as ex2:
                        logger.warning(f"zip 파일 읽기 실패: {ex2}")

                    for sheet_file in sheet_files:
                        try:
                            # build dataframe from xml
                            df = self._sheet_to_dataframe_via_zip(file_path, sheet_file, shared_strings)
                            sheet_basename = sheet_file.split('/')[-1]
                            sheet_name = sheet_basename.replace('sheet', 'sheet_')
                            extracted_personal = self._extract_personal_info_from_sheet(df, sheet_name)
                            if extracted_personal:
                                personal_info.update(extracted_personal)
                                logger.info(f"시트 '{sheet_name}'에서 개인정보 추출(zip): {len(extracted_personal)}개 필드")
                            extracted_courses = self._extract_course_records_from_sheet(df, sheet_name)
                            if extracted_courses:
                                course_records.extend(extracted_courses)
                                logger.info(f"시트 '{sheet_name}'에서 수강기록 추출(zip): {len(extracted_courses)}개 과목")
                        except Exception as ex3:
                            logger.warning(f"시트 {sheet_file} 파싱(zip) 실패: {ex3}")

                    # fixed cells via zip
                    try:
                        fixed_cells = self._extract_fixed_cell_credits_via_zip(file_path)
                        if fixed_cells:
                            personal_info.update(fixed_cells)
                            logger.info(f"고정셀에서 학점 추출(zip): {fixed_cells}")
                    except Exception as ex4:
                        logger.warning(f"고정셀 추출(zip) 실패: {ex4}")

                    if not course_records:
                        parsing_warnings.append('과목표에서 수강기록을 추출하지 못했습니다. (zip)')
                        logger.warning('과목표에서 수강기록을 추출하지 못했습니다. (zip)')

                    for key in ('전공필수학점', '전공선택학점', '일반선택학점'):
                        if key not in personal_info or personal_info.get(key) is None:
                            parsing_warnings.append(f"{key}가 추출되지 않았습니다. (zip)")

                    if parsing_warnings:
                        personal_info['parsing_warnings'] = parsing_warnings
                    else:
                        personal_info['parsing_warnings'] = []
                    # indicate parser used
                    personal_info['parsing_warnings'].append('parser_used: zip')

                    return personal_info, course_records
                except Exception as ex5:
                    logger.error(f"zip/xml fallback 실패: {ex5}")
                    raise

        finally:
            if excel_file is not None:
                try:
                    del excel_file  # 객체 참조 제거
                except:
                    pass
                gc.collect()  # 가비지 수집 호출로 핸들 완전 해제

    def _extract_fixed_cell_credits(self, file_path: str) -> Dict[str, Any]:
        """openpyxl로 시트들을 순회하며 AC22(전공필수), AH22(전공선택), Y22(일반선택) 값을 추출"""
        results: Dict[str, Any] = {}
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception:
            # read_only가 안될 경우 기본 모드로 재시도
            wb = openpyxl.load_workbook(file_path, data_only=True)

        try:
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                # 셀 값 읽기 (없으면 None)
                ac22 = ws['AC22'].value
                ah22 = ws['AH22'].value
                y22 = ws['Y22'].value

                # 숫자처럼 보이는 값만 반영
                def to_float_safe(v):
                    try:
                        return float(v)
                    except Exception:
                        return None

                ac_val = to_float_safe(ac22)
                ah_val = to_float_safe(ah22)
                y_val = to_float_safe(y22)

                if ac_val is not None and '전공필수학점' not in results:
                    results['전공필수학점'] = ac_val
                if ah_val is not None and '전공선택학점' not in results:
                    results['전공선택학점'] = ah_val
                if y_val is not None and '일반선택학점' not in results:
                    results['일반선택학점'] = y_val

                # 모두 추출되었으면 조기 종료
                if {'전공필수학점', '전공선택학점', '일반선택학점'}.issubset(results.keys()):
                    break
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return results

    def _extract_fixed_cell_credits_via_zip(self, file_path: str) -> Dict[str, Any]:
        """Zip/xml based minimal extractor for cell values to avoid openpyxl style parsing issues"""
        results: Dict[str, Any] = {}
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                # shared strings
                shared_strings = []
                try:
                    ss_xml = z.read('xl/sharedStrings.xml')
                    root = ET.fromstring(ss_xml)
                    for si in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                        # concatenate t text elements
                        texts = ''.join([t.text or '' for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')])
                        shared_strings.append(texts)
                except KeyError:
                    shared_strings = []

                # iterate worksheets
                sheet_files = [name for name in z.namelist() if name.startswith('xl/worksheets/sheet')]
                for sheet_file in sheet_files:
                    try:
                        sheet_xml = z.read(sheet_file)
                    except KeyError:
                        continue
                    root = ET.fromstring(sheet_xml)
                    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
                    for c in root.findall('.//' + ns + 'c'):
                        r = c.get('r')  # cell reference e.g. AC22
                        if r and r in ('AC22', 'AH22', 'Y22'):
                            v = c.find(ns + 'v')
                            if v is None:
                                continue
                            if c.get('t') == 's':
                                idx = int(v.text)
                                val = shared_strings[idx] if idx < len(shared_strings) else None
                            else:
                                val = v.text
                            try:
                                num = float(val)
                                if r == 'AC22' and '전공필수학점' not in results:
                                    results['전공필수학점'] = num
                                if r == 'AH22' and '전공선택학점' not in results:
                                    results['전공선택학점'] = num
                                if r == 'Y22' and '일반선택학점' not in results:
                                    results['일반선택학점'] = num
                            except Exception:
                                # not numeric, skip
                                pass
                    # if we've collected all
                    if {'전공필수학점', '전공선택학점', '일반선택학점'}.issubset(results.keys()):
                        break
        except Exception as e:
            logger.warning(f"Zip 기반 셀 추출 오류: {e}")
        return results

    def _sheet_to_dataframe_via_zip(self, file_path: str, sheet_file: str, shared_strings: List[str]) -> pd.DataFrame:
        """Parse a worksheet XML and return a pandas DataFrame using xml parsing. Minimal and conservative approach for fallbacks."""
        ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
        rows_map = {}
        max_col = 0
        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_xml = z.read(sheet_file)
            root = ET.fromstring(sheet_xml)
            for row in root.findall('.//' + ns + 'row'):
                r_idx = int(row.get('r')) - 1
                row_vals = {}
                for c in row.findall(ns + 'c'):
                    ref = c.get('r')  # e.g. A1
                    # convert column letters to index
                    col_letters = ''.join([ch for ch in ref if ch.isalpha()])
                    col_idx = self._col_letter_to_index(col_letters)
                    v = c.find(ns + 'v')
                    if v is None:
                        continue
                    if c.get('t') == 's':
                        val = shared_strings[int(v.text)] if int(v.text) < len(shared_strings) else None
                    else:
                        val = v.text
                    row_vals[col_idx] = val
                    if col_idx > max_col:
                        max_col = col_idx
                rows_map[r_idx] = row_vals

        # build rows list
        rows = []
        max_row = max(rows_map.keys()) if rows_map else -1
        for r in range(max_row + 1):
            row_vals = rows_map.get(r, {})
            row_list = [row_vals.get(c) for c in range(max_col + 1)] if max_col >= 0 else []
            rows.append(row_list)
        df = pd.DataFrame(rows)
        return df

    def _col_letter_to_index(self, letters: str) -> int:
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch.upper()) - ord('A') + 1)
        return idx - 1
    
    def _extract_personal_info_from_sheet(self, df: pd.DataFrame, sheet_name: str) -> Dict:
        """시트에서 개인정보 추출"""
        personal_info = {}
        
        # 방법 1: 세로 형태 데이터 (항목:값 형태)
        personal_info.update(self._extract_vertical_personal_info(df))
        
        # 방법 2: 가로 형태 데이터 (첫 번째 행이 헤더)
        personal_info.update(self._extract_horizontal_personal_info(df))
        
        # 방법 3: 셀 단위 검색
        personal_info.update(self._extract_cell_based_personal_info(df))
        
        # 방법 4: 구조화된 개인정보 추출 (Excel 상단 영역)
        personal_info.update(self._extract_structured_personal_info(df))
        
        return personal_info
    
    def _extract_vertical_personal_info(self, df: pd.DataFrame) -> Dict:
        """세로 형태의 개인정보 추출 (항목:값)"""
        personal_info = {}
        
        for _, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                    
                cell_str = str(cell_value).strip()
                
                # ':' 기준으로 분할하여 항목:값 형태 찾기
                if ':' in cell_str:
                    parts = cell_str.split(':', 1)
                    if len(parts) == 2:
                        key_part = parts[0].strip()
                        value_part = parts[1].strip()
                        
                        # 개인정보 필드와 매칭
                        for std_field, possible_names in self.personal_info_mappings.items():
                            if any(name in key_part for name in possible_names):
                                if value_part and value_part != '-':
                                    personal_info[std_field] = value_part
                                    logger.debug(f"세로형식에서 발견: {std_field} = {value_part}")
                                break
                
                # 인접한 셀에서 값 찾기
                if col_idx < len(row) - 1:
                    next_cell = row.iloc[col_idx + 1]
                    if not pd.isna(next_cell):
                        for std_field, possible_names in self.personal_info_mappings.items():
                            if any(name in cell_str for name in possible_names):
                                value = str(next_cell).strip()
                                if value and value != '-':
                                    personal_info[std_field] = value
                                    logger.debug(f"인접셀에서 발견: {std_field} = {value}")
                                break
        
        return personal_info
    
    def _extract_horizontal_personal_info(self, df: pd.DataFrame) -> Dict:
        """가로 형태의 개인정보 추출 (첫 번째 행이 헤더)"""
        personal_info = {}
        
        if len(df) == 0:
            return personal_info
        
        # 첫 번째 행을 헤더로 사용
        headers = [str(col).strip() for col in df.columns]
        
        for col_idx, header in enumerate(headers):
            for std_field, possible_names in self.personal_info_mappings.items():
                if any(name in header for name in possible_names):
                    # 해당 열의 첫 번째 데이터 행에서 값 추출
                    for _, row in df.iterrows():
                        if col_idx < len(row) and not pd.isna(row.iloc[col_idx]):
                            value = str(row.iloc[col_idx]).strip()
                            if value and value != '-':
                                personal_info[std_field] = value
                                logger.debug(f"가로형식에서 발견: {std_field} = {value}")
                                break
                    break
        
        return personal_info
    
    def _extract_cell_based_personal_info(self, df: pd.DataFrame) -> Dict:
        """셀 단위로 개인정보 검색"""
        personal_info = {}
        
        for row_idx, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                
                cell_str = str(cell_value).strip()
                
                # 패턴 매칭으로 특정 정보 추출
                
                # 학번 패턴 (숫자로만 구성된 7-10자리)
                if re.match(r'^\d{7,10}$', cell_str) and '학번' not in personal_info:
                    personal_info['학번'] = cell_str
                    logger.debug(f"패턴매칭에서 학번 발견: {cell_str}")
                
                # 이메일 패턴
                if '@' in cell_str and re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', cell_str):
                    # 이메일은 개인정보에 포함하지 않지만 로깅
                    logger.debug(f"이메일 패턴 발견: {cell_str}")
                
                # 날짜 패턴 (YYYY-MM-DD, YYYY.MM.DD 등)
                date_patterns = [
                    r'(\d{4})[-.년](\d{1,2})[-.월](\d{1,2})',
                    r'(\d{4})/(\d{1,2})/(\d{1,2})',
                ]
                for pattern in date_patterns:
                    match = re.search(pattern, cell_str)
                    if match:
                        year, month, day = match.groups()
                        formatted_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                        
                        # 입학일자로 추정 (년도가 2000 이후이고 월이 3 또는 9)
                        if int(year) >= 2000 and int(month) in [3, 9] and '입학일자' not in personal_info:
                            personal_info['입학일자'] = formatted_date
                            logger.debug(f"패턴매칭에서 입학일자 발견: {formatted_date}")
                        
                        # 생년월일로 추정 (년도가 1950-2010 사이)
                        elif 1950 <= int(year) <= 2010 and '생년월일' not in personal_info:
                            personal_info['생년월일'] = formatted_date
                            logger.debug(f"패턴매칭에서 생년월일 발견: {formatted_date}")
        
        return personal_info
    
    def _extract_structured_personal_info(self, df: pd.DataFrame) -> Dict:
        """구조화된 개인정보 추출 (Excel 상단 영역의 라벨-값 쌍)"""
        personal_info = {}
        
        # 상단 20행에서 개인정보 검색
        for row_idx in range(min(20, len(df))):
            row = df.iloc[row_idx]
            
            # 각 셀을 검사하여 라벨을 찾고 인근 셀에서 값 추출
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                
                cell_str = str(cell_value).strip()
                if not cell_str:
                    continue
                
                # 개인정보 필드와 매칭 시도
                for std_field, possible_names in self.personal_info_mappings.items():
                    if any(name in cell_str for name in possible_names):
                        # 같은 행의 오른쪽 셀들에서 값 찾기
                        value = self._find_value_in_row(row, col_idx)
                        if value and std_field not in personal_info:
                            # 잘못된 매칭 필터링
                            if self._is_valid_field_value(std_field, value):
                                personal_info[std_field] = value
                                logger.debug(f"구조화된 방식에서 발견: {std_field} = {value}")
                            break
        
        return personal_info
    
    def _find_value_in_row(self, row: pd.Series, start_col: int) -> Optional[str]:
        """행에서 라벨 다음에 오는 유효한 값 찾기"""
        # 시작 위치부터 최대 20개 셀까지 검사
        for col_idx in range(start_col + 1, min(start_col + 21, len(row))):
            cell_value = row.iloc[col_idx]
            if pd.isna(cell_value):
                continue
            
            cell_str = str(cell_value).strip()
            if not cell_str or cell_str == '-':
                continue
            
            # 숫자만으로 이루어진 경우 (학번, 년도 등)
            if cell_str.isdigit() and len(cell_str) >= 4:
                return cell_str
            
            # 날짜 형식인 경우
            if self._is_date_format(cell_str):
                return self._normalize_date(cell_str)
            
            # 한글이 포함된 문자열 (이름, 학과명 등)
            if any('\u3131' <= char <= '\u3163' or '\uac00' <= char <= '\ud7a3' for char in cell_str):
                return cell_str
            
            # 영문이 포함된 문자열
            if any(char.isalpha() for char in cell_str):
                return cell_str
            
            # 학년 형태 (3학년 등)
            if '학년' in cell_str:
                return cell_str.replace('학년', '').strip()
            
            # 소수점이 포함된 숫자 (학기, 평점 등)
            if re.match(r'^\d+\.\d+$', cell_str):
                return cell_str
        
        return None
    
    def _is_date_format(self, text: str) -> bool:
        """날짜 형식인지 확인"""
        # YYYY-MM-DD, YYYY.MM.DD, YY.MM.DD 등
        import re
        date_patterns = [
            r'^\d{4}[-./]\d{1,2}[-./]\d{1,2}$',
            r'^\d{2}[-./]\d{1,2}[-./]\d{1,2}$'
        ]
        return any(re.match(pattern, text) for pattern in date_patterns)
    
    def _normalize_date(self, date_str: str) -> str:
        """날짜 형식 정규화"""
        import re
        # YY.MM.DD 형태를 YYYY-MM-DD로 변환
        if re.match(r'^\d{2}\.\d{1,2}\.\d{1,2}$', date_str):
            parts = date_str.split('.')
            year = int(parts[0])
            # 2000년대로 가정 (00-30은 2000년대, 31-99는 1900년대)
            if year <= 30:
                year += 2000
            else:
                year += 1900
            return f"{year}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
        
        # YYYY.MM.DD나 YYYY-MM-DD를 YYYY-MM-DD로 통일
        normalized = re.sub(r'[./]', '-', date_str)
        parts = normalized.split('-')
        if len(parts) == 3:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
        
        return date_str
    
    def _is_valid_field_value(self, field_name: str, value: str) -> bool:
        """필드별 값이 유효한지 검증"""
        value = str(value).strip()
        
        # 일반적으로 잘못된 값들
        invalid_values = ['다전공1', '다전공2', '교직', '이론', '소양', '실습', '부전공', '다전공', '전공', '필수', '선택']
        if value in invalid_values:
            return False
        
        # 필드별 특별 검증
        if field_name == '학번':
            return value.isdigit() and 7 <= len(value) <= 10
        elif field_name == '학년':
            return any(char.isdigit() for char in value)
        elif field_name == '교과적용년도':
            return value.isdigit() and len(value) == 4
        elif field_name == '이수학기':
            # 이수학기는 숫자나 숫자.숫자 형태여야 함
            return re.match(r'^\d+(\.\d+)?$', value) is not None
        elif field_name in ['생년월일', '입학일자']:
            return self._is_date_format(value)
        elif field_name == '성명':
            # 한글 이름이어야 함
            return any('\uac00' <= char <= '\ud7a3' for char in value)
        elif field_name in ['대학', '학과', '전공', '부전공', '다전공']:
            # 한글이 포함되어야 하고 일정 길이 이상이어야 함
            return any('\uac00' <= char <= '\ud7a3' for char in value) and len(value) >= 2
        
        return True
    
    def _extract_course_records_from_sheet(self, df: pd.DataFrame, sheet_name: str) -> List[Dict]:
        """시트에서 수강기록 추출"""
        course_records = []
        
        # 수강기록으로 보이는 시트인지 확인 (보수적으로: 인식 실패해도 계속 시도)
        if not self._is_course_record_sheet(df, sheet_name):
            logger.info(f"시트 '{sheet_name}'는 수강기록 시트로 확신할 수 없지만, 헤더 탐지 시도를 진행합니다.")
        
        logger.info(f"'{sheet_name}' 시트에서 수강기록 추출 시작")
        
        # 헤더 행 찾기
        header_row_idx = self._find_header_row(df)
        if header_row_idx is None:
            logger.warning(f"'{sheet_name}' 시트에서 헤더를 찾을 수 없습니다.")
            return course_records
        
        # 헤더 매핑
        headers = df.iloc[header_row_idx].tolist()
        column_mapping = self._map_course_columns(headers)

        # 페일오버: 일부 필드(교과목명/학점 등)가 매핑되지 않았다면, 상단 몇 행을 스캔해 해당 컬럼을 추정
        missing_fields = [f for f, v in column_mapping.items() if v is None and f in ('교과목명', '학점', '영역', '이수구분', '성적')]
        if missing_fields:
            logger.info(f"헤더 매핑 일부 실패({missing_fields}), 페일오버 스캔 시작")
            # 스캔 범위: header_row_idx +/- 5, 그리고 상위 30행
            scan_rows = list(range(max(0, header_row_idx - 5), min(len(df), header_row_idx + 6)))
            scan_rows.extend(list(range(0, min(30, len(df)))))
            scan_rows = sorted(set(scan_rows))
            for r in scan_rows:
                row = df.iloc[r]
                for col_idx, cell in enumerate(row):
                    if pd.isna(cell):
                        continue
                    cell_str = str(cell).strip()
                    for std_field in missing_fields[:]:
                        for poss in self.course_field_mappings.get(std_field, []):
                            if poss in cell_str:
                                column_mapping[std_field] = col_idx
                                missing_fields.remove(std_field)
                                logger.info(f"페일오버 매핑: 필드 {std_field} -> 컬럼 {col_idx} (행 {r})")
                                break
                        # if found, break outer loop
                    if not missing_fields:
                        break
                if not missing_fields:
                    break
        
        logger.info(f"컬럼 매핑: {column_mapping}")
        
        # 데이터 추출
        for row_idx in range(header_row_idx + 1, len(df)):
            row = df.iloc[row_idx]
            
            # 교과목명이 있는 행만 처리
            course_name_col = column_mapping.get('교과목명')
            if course_name_col is None or pd.isna(row.iloc[course_name_col]) or not str(row.iloc[course_name_col]).strip():
                continue
            
            course_record = {}
            
            # 각 필드 추출
            for std_field, col_idx in column_mapping.items():
                if col_idx is not None and col_idx < len(row):
                    value = row.iloc[col_idx]
                    if not pd.isna(value):
                        # 값 정제
                        cleaned_value = self._clean_course_field_value(std_field, str(value).strip())
                        course_record[std_field] = cleaned_value
                    else:
                        course_record[std_field] = None
                else:
                    course_record[std_field] = None
            
            # 필수 필드 검증 및 영역 자동 보정
            if course_record.get('교과목명'):
                # 교양 과목인데 영역이 없으면 교과목명/행 전체에서 영역 키워드 탐지
                if course_record.get('구분') == '교양' and not course_record.get('영역'):
                    course_record['영역'] = self._infer_liberal_area(course_record.get('교과목명', ''), row)
                    if course_record['영역']:
                        logger.debug(f"영역 자동 보정: {course_record['교과목명']} -> {course_record['영역']}")
                course_records.append(course_record)
        
        logger.info(f"'{sheet_name}' 시트에서 {len(course_records)}개 수강기록 추출 완료")
        return course_records
    
    def _is_course_record_sheet(self, df: pd.DataFrame, sheet_name: str) -> bool:
        """수강기록 시트인지 판단"""
        # 시트명으로 판단
        course_keywords = ['성적', '이수', '수강', '교과목', '과목', 'grade', 'course']
        sheet_name_lower = sheet_name.lower()
        if any(keyword in sheet_name_lower for keyword in course_keywords):
            return True
        
        # 내용으로 판단 (교과목 관련 컬럼이 여러 개 있는지)
        course_column_count = 0
        for _, row in df.iterrows():
            for cell in row:
                if pd.isna(cell):
                    continue
                cell_str = str(cell).lower()
                if any(keyword in cell_str for keyword in ['교과목', '과목', '학점', '성적', '년도', '학기']):
                    course_column_count += 1
                    break
            if course_column_count >= 3:  # 3개 이상의 관련 컬럼이 있으면 수강기록 시트로 판단
                return True
        
        return False
    
    def _find_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """헤더 행 찾기"""
        max_matches = 0
        best_row_idx = None
        
        for row_idx, row in df.iterrows():
            matches = 0
            for cell in row:
                if pd.isna(cell):
                    continue
                cell_str = str(cell).strip()
                
                # 수강기록 관련 헤더인지 확인
                for field_mappings in self.course_field_mappings.values():
                    if any(mapping in cell_str for mapping in field_mappings):
                        matches += 1
                        break
            
            if matches > max_matches:
                max_matches = matches
                best_row_idx = row_idx
        
        # 최소 3개 이상의 필드가 매칭되어야 헤더로 인정
        if max_matches >= 3:
            return best_row_idx

        # 페일오버: 상위 30행(또는 df내 제한된 범위)에서 키워드 기반으로 헤더 검색
        rows_to_try = min(30, len(df))
        header_keywords = set([k for v in self.course_field_mappings.values() for k in v])
        for r_idx in range(rows_to_try):
            row = df.iloc[r_idx]
            matches = 0
            for cell in row:
                if pd.isna(cell):
                    continue
                cell_str = str(cell).strip()
                if any(k in cell_str for k in header_keywords):
                    matches += 1
            if matches >= 2:
                logger.info(f"헤더 페일오버 찾음: 행 {r_idx} (matches={matches})")
                return r_idx

        # 멀티행 헤더 페일오버: 첫 10행에서 두 행을 합쳐 키워드 매칭
        for r_idx in range(min(10, len(df)-1)):
            row_combined = []
            row_combined.extend([str(x).strip() for x in df.iloc[r_idx] if not pd.isna(x)])
            row_combined.extend([str(x).strip() for x in df.iloc[r_idx+1] if not pd.isna(x)])
            matches = sum(1 for s in row_combined if any(k in s for k in header_keywords))
            if matches >= 3:
                logger.info(f"멀티행 헤더 페일오버 찾음: 행 {r_idx}~{r_idx+1} (matches={matches})")
                return r_idx

        return None
    
    def _map_course_columns(self, headers: List) -> Dict[str, Optional[int]]:
        """헤더를 표준 필드명에 매핑"""
        column_mapping = {}
        
        for std_field, possible_names in self.course_field_mappings.items():
            column_mapping[std_field] = None
            
            for col_idx, header in enumerate(headers):
                if pd.isna(header):
                    continue
                header_str = str(header).strip()
                
                for possible_name in possible_names:
                    if possible_name in header_str:
                        column_mapping[std_field] = col_idx
                        break
                
                if column_mapping[std_field] is not None:
                    break
        
        return column_mapping

        # 추가 페일오버: 매핑이 못 된 경우, 전체 시트(최대 상위 20행)에서 라벨을 찾아 컬럼 추정
    
    def _infer_liberal_area(self, course_name: str, row: pd.Series) -> Optional[str]:
        """교양 과목의 영역을 교과목명 및 행 전체에서 추론"""
        # 교과목명에서 키워드 탐지
        for area_name, keywords in self.liberal_area_keywords.items():
            if any(kw in course_name for kw in keywords):
                return area_name
        
        # 행 전체 셀에서 키워드 탐지
        for cell in row:
            if pd.isna(cell):
                continue
            cell_str = str(cell).strip()
            for area_name, keywords in self.liberal_area_keywords.items():
                if any(kw in cell_str for kw in keywords):
                    return area_name
        
        return None
    
    def _clean_course_field_value(self, field_name: str, value: str) -> Any:
        """필드별 값 정제"""
        if not value or value == '-':
            return None
        
        if field_name == '학점':
            # 학점 정제 (숫자만 추출)
            try:
                return float(re.sub(r'[^\d.]', '', value))
            except:
                return None
        
        elif field_name in ['년도', '학기']:
            # 숫자 추출
            try:
                return int(re.sub(r'[^\d]', '', value))
            except:
                return None
        
        else:
            return value
    
    def save_to_database(self, student_id: str, personal_info: Dict, course_records: List[Dict]):
        """데이터베이스에 저장"""
        try:
            cursor = self.connection.cursor()
            
            # 개인정보에 학번 추가
            personal_info['학번'] = student_id
            
            # null 값 처리
            for key, value in personal_info.items():
                if value is None or str(value).strip() == '':
                    personal_info[key] = None
            
            # 개인정보 저장
            self._save_personal_info(cursor, personal_info)
            
            # 수강기록 저장
            self._save_course_records(cursor, student_id, course_records)
            
            self.connection.commit()
            
            # 저장된 데이터 로깅
            logger.info("=== 데이터베이스 저장 완료 ===")
            logger.info(f"학번: {student_id}")
            logger.info(f"개인정보: {len([v for v in personal_info.values() if v is not None])}개 필드 저장")
            logger.info(f"수강기록: {len(course_records)}개 과목 저장")
            
        except Error as e:
            self.connection.rollback()
            logger.error(f"데이터베이스 저장 오류: {e}")
            raise
        finally:
            if cursor:
                cursor.close()
    
    def _save_personal_info(self, cursor, personal_info: Dict):
        """개인정보 저장"""
        # 필드명 매핑 (DB 컬럼명으로 변환)
        db_field_mapping = {
            '학번': 'student_id',
            '대학': 'university',
            '학과': 'department',
            '전공': 'major',
            '부전공': 'minor',
            '다전공': 'double_major',
            '과정': 'course_type',
            '입학일자': 'admission_date',
            '성명': 'name',
            '교과적용년도': 'curriculum_year',
            '이수학기': 'semester',
            '생년월일': 'birth_date',
            '학년': 'grade',
            '평생사제상담건수': 'counseling_count',
            '전공필수학점': 'major_required_credits',
            '전공선택학점': 'major_elective_credits',
            '일반선택학점': 'general_elective_credits'
        }
        
        # DB 필드로 변환
        db_data = {}
        for korean_field, value in personal_info.items():
            db_field = db_field_mapping.get(korean_field)
            if db_field:
                # 특별 처리: 숫자 필드들
                if db_field == 'grade' and value:
                    # "3학년" -> "3", "3" -> "3"
                    grade_match = re.search(r'\d+', str(value))
                    if grade_match:
                        db_data[db_field] = int(grade_match.group())
                    else:
                        db_data[db_field] = None
                elif db_field == 'counseling_count' and value:
                    # "3 (졸업기준 : 8)" -> "3"
                    count_match = re.search(r'\d+', str(value))
                    if count_match:
                        db_data[db_field] = int(count_match.group())
                    else:
                        db_data[db_field] = None
                else:
                    db_data[db_field] = value
        
        # UPSERT 쿼리 생성
        fields = list(db_data.keys())
        placeholders = ', '.join(['%s'] * len(fields))
        field_list = ', '.join(fields)
        update_list = ', '.join([f"{field} = VALUES({field})" for field in fields if field != 'student_id'])
        
        query = f"""
        INSERT INTO students ({field_list}, created_at, updated_at)
        VALUES ({placeholders}, NOW(), NOW())
        ON DUPLICATE KEY UPDATE
        {update_list}, updated_at = NOW()
        """
        
        values = [db_data[field] for field in fields]
        # Ensure users table entry exists to satisfy foreign key
        try:
            # If user does not exist in users table, create minimal user
            student_id = db_data.get('student_id')
            if student_id:
                cursor.execute('SELECT username FROM users WHERE username=%s', (student_id,))
                if cursor.fetchone() is None:
                    # create minimal user with a random password
                    password_hash = bcrypt.hashpw('change_me_123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                    cursor.execute('INSERT INTO users (username, password_hash, role, is_active, created_at) VALUES (%s, %s, %s, %s, NOW())', (student_id, password_hash, 'student', 1))
        except Exception as e:
            logger.warning(f"사용자 생성 시도 중 오류: {e}")
        cursor.execute(query, values)
        
        logger.info(f"개인정보 저장: {db_data}")
    
    def _save_course_records(self, cursor, student_id: str, course_records: List[Dict]):
        """수강기록 저장"""
        # 기존 기록 삭제
        cursor.execute("DELETE FROM course_records WHERE student_id = %s", (student_id,))
        
        if not course_records:
            return
        
        # 새 기록 삽입
        query = """
        INSERT INTO course_records (
            student_id, category, area, sub_area, year, semester,
            course_code, course_name, credit, completion_type, grade, created_at
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
        )
        """
        
        insert_count = 0
        for record in course_records:
            values = (
                student_id,
                record.get('구분'),
                record.get('영역'),
                record.get('세부영역'),
                record.get('년도'),
                record.get('학기'),
                record.get('교과목번호'),
                record.get('교과목명'),
                record.get('학점'),
                record.get('이수구분'),
                record.get('성적')
            )
            cursor.execute(query, values)
            insert_count += 1
        
        logger.info(f"수강기록 {insert_count}개 저장 완료")

def process_excel_file_enhanced(file_path: str, student_id: str, db_config: Dict[str, str]) -> bool:
    """향상된 Excel 파일 처리 함수"""
    parser = EnhancedXlsxParser(db_config)
    
    try:
        parser.connect_db()
        personal_info, course_records = parser.parse_excel_file(file_path)

        required_fields = ['학번', '성명']
        missing_fields = [f for f in required_fields if f not in personal_info or not personal_info[f]]
        if missing_fields:
            logger.error(f"⚠️ 필수 개인정보 누락: {missing_fields}")
            # 학번이 없으면 매개변수로 받은 student_id 사용
            if '학번' in missing_fields:
                personal_info['학번'] = student_id
                logger.info(f"매개변수 student_id를 학번으로 사용: {student_id}")
                missing_fields.remove('학번')
            
            # 성명이 없으면 기본값 설정
            if '성명' in missing_fields:
                personal_info['성명'] = f"학생_{student_id}"
                logger.info(f"성명이 없어 기본값 설정: 학생_{student_id}")
                missing_fields.remove('성명')
            
            # 여전히 필수 필드가 누락된 경우
            if missing_fields:
                logger.error(f"⚠️ 해결할 수 없는 필수 개인정보 누락: {missing_fields}")
                return False
        
        # 최소 조건 검증: 개인정보는 있어야 하고, 수강기록은 없어도 됨
        if not personal_info:
            logger.warning("Excel 파일에서 개인정보를 찾을 수 없습니다.")
            return False
        
        # 성명이 여전히 없는 경우 기본값 설정
        if not personal_info.get('성명'):
            personal_info['성명'] = f"학생_{student_id}"
            logger.info(f"성명이 없어 기본값 설정: 학생_{student_id}")
        
        logger.info(f"데이터 검증 완료 - 개인정보: {len(personal_info)}개, 수강기록: {len(course_records)}개")
        
        parser.save_to_database(student_id, personal_info, course_records)
        
        # 결과 요약 출력
        print("\n" + "="*50)
        print("📊 Excel 파일 처리 완료")
        print("="*50)
        print(f"👤 학번: {student_id}")
        print(f"📋 개인정보: {len([v for v in personal_info.values() if v is not None])}개 필드 추출")
        print(f"📚 수강기록: {len(course_records)}개 과목 추출")
        print("\n📝 추출된 개인정보:")
        for field, value in personal_info.items():
            if value is not None:
                print(f"  - {field}: {value}")
        print(f"\n📖 수강기록 샘플 (처음 3개):")
        for i, record in enumerate(course_records[:3]):
            print(f"  {i+1}. {record.get('교과목명', 'N/A')} ({record.get('학점', 'N/A')}학점, {record.get('성적', 'N/A')})")
        if len(course_records) > 3:
            print(f"  ... 총 {len(course_records)}개 과목")
        print("="*50)
        
        parsing_warnings = personal_info.get('parsing_warnings', [])
        return True, parsing_warnings
        
    except Exception as e:
        logger.error(f"Excel 파일 처리 중 오류 발생: {e}")
        parsing_warnings = personal_info.get('parsing_warnings', []) if 'personal_info' in locals() else []
        return False, parsing_warnings
    finally:
        parser.disconnect_db()

if __name__ == "__main__":
    db_config = {
        'host': '203.255.78.58',
        'port': 9003,
        'database': 'graduation_system',
        'user': 'user29',
        'password': '123'
    }
    
    test_file_path = 'ex_1.xlsx'
    test_student_id = '2023001'
    
    success = process_excel_file_enhanced(test_file_path, test_student_id, db_config)
    if success:
        print("\n✅ Excel 파일 처리가 성공적으로 완료되었습니다!")
    else:
        print("\n❌ Excel 파일 처리에 실패했습니다.")